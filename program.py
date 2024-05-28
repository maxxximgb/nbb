import os
import time

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import sys
import tkinter as tk
from tkinter import messagebox, Button, filedialog, simpledialog
from tkinter.scrolledtext import ScrolledText
import shutil
import psutil
import pandas as pd


class Infopovod:
    def __init__(self, data, styles):
        self.data = data
        self.styles = styles


class Ministerstvo:
    def __init__(self, name, filename):
        self.name = name
        self.filename = filename
        self.infopovody = []


def sort_keys_by_importance(key):
    order_of_importance = ['Министерство', 'Комитет', 'Служба', 'Управление', 'АМС']
    for i, category in enumerate(order_of_importance):
        if category in key:
            return i
    return len(order_of_importance)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Обработка данных')
        self.geometry('800x700')
        self.initUI()

    def initUI(self):
        self.output_text = ScrolledText(self)
        self.output_text.pack(fill=tk.BOTH, expand=True)
        self.button = Button(self, text='Обработать данные', command=self.process_data)
        self.button.pack()

    def process_data(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            ministerstva = self.process_files(folder_path)
            self.create_excel_file(ministerstva, folder_path)

        else:
            messagebox.showwarning("Ошибка", "Инфоповоды не загружены или не обнаружены.")
            sys.exit(0)

    def process_files(self, folder_path):
        self.output_text.insert(tk.END, 'Начинается обработка файлов.\n')
        ministerstva = {}
        try:
            if self.is_directory_empty(folder_path):
                messagebox.showwarning("Ошибка", "Инфоповоды не загружены или не обнаружены.\n")
                os.abort()
        except:
            messagebox.showwarning("Ошибка", "Инфоповоды не загружены или не обнаружены.")
            sys.exit(0)
        else:
            self.output_text.insert(tk.END, "Инфоповоды найдены!\n")
        for filename in os.listdir(folder_path):
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                if filename.endswith('.xls'):
                    filename = self.convert_xls_to_xlsx(folder_path, filename)
                book = load_workbook(os.path.join(folder_path, filename))
                sheet = book.active
                row = 2
                while sheet[row][6].value is not None:
                    data_row = [cell.value for cell in sheet[row]]
                    styles_row = [cell.font for cell in sheet[row]]
                    ministerstvo_name = data_row[6]
                    infopovod = Infopovod(data_row, styles_row)
                    if ministerstvo_name not in ministerstva:
                        ministerstva[ministerstvo_name] = Ministerstvo(ministerstvo_name, filename)
                    ministerstva[ministerstvo_name].infopovody.append(infopovod)
                    self.output_text.insert(tk.END,
                                            'Инфоповод министерства {} с номером строки {} был обработан\n'.format(
                                                ministerstvo_name, row))
                    # Добавление обводки к 8-му столбцу
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    # Создание восьмого столбца, если он не существует
                    if len(sheet[row]) < 8:
                        sheet.cell(row=row, column=8, value='')
                    sheet[row][7].border = thin_border
                    row += 1
                # Сохранение изменений в книге
                book.save(os.path.join(folder_path, filename))
        return ministerstva

    def create_excel_file(self, ministerstva, folder_path):
        self.output_text.insert(tk.END, 'Начата генерация выходного файла\n')
        wb = Workbook()
        ws = wb.active
        row = 1
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        date = simpledialog.askstring("Введите дату", "Введите дату")
        date_cell = ws.cell(row=row, column=2, value=date)
        date_cell.number_format = 'DD/MM/YYYY'
        date_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        date_cell.font = Font(name='Times New Roman', bold=True, size=16, color="0000FF")  # Добавлено: color="0000FF"
        date_cell.alignment = Alignment(horizontal='left')
        row += 1
        print(sorted(ministerstva.keys()))
        for ministerstvo_name in sorted(ministerstva.keys(), key=sort_keys_by_importance):
            ministerstvo = ministerstva[ministerstvo_name]
            for infopovod in ministerstvo.infopovody:
                for j, (data, style) in enumerate(zip(infopovod.data, infopovod.styles)):
                    cell = ws.cell(row=row, column=j + 1, value=data)
                    cell.font = Font(name='Times New Roman', bold=style.bold, italic=style.italic)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    if j < 8:  # Обновлено: добавлено условие для обводки
                        cell.border = thin_border
                    if j == 1:
                        cell.number_format = 'DD/MM/YYYY'
                    if j == 7:  # Добавлено: условие для восьмого столбца
                        cell.value = ""
                    book = load_workbook(os.path.join(folder_path, ministerstvo.filename))
                    sheet = book.active
                    ws.column_dimensions[get_column_letter(j + 1)].width = sheet.column_dimensions[
                        get_column_letter(j + 1)].width
                    ws.row_dimensions[row].height = None
                row += 1
        self.output_text.insert(tk.END, 'Файл успешно сгенерирован и сохранен\n')
        reply = messagebox.askquestion('Удалить старые инфоповоды из папки?', 'Удалить старые инфоповоды из папки?',
                                       icon='warning')
        if reply == 'yes':
            self.output_text.insert(tk.END, 'Инфоповоды удалены\n')
            self.delete_all_files_in_directory(folder_path)
        try:
            wb.save('output.xlsx')
        except PermissionError:
            uc = messagebox.askyesno('Внимание!', 'Необходимо закрыть excel, закрыть?',
                                     icon='warning')
            if uc == 1:
                for proc in psutil.process_iter():
                    if proc.name() == 'EXCEL.EXE':
                        proc.kill()
                        time.sleep(3)
        wb.save('output.xlsx')
        os.system('start output.xlsx')
        response = messagebox.askyesno("Внимание",
                                       "Сгенерированная таблица соответствуеет ожиданиям?")
        if response != 1:
            for proc in psutil.process_iter():
                if proc.name() == 'EXCEL.EXE':
                    proc.kill()
                    sys.exit(0)
        else:
            for proc in psutil.process_iter():
                if proc.name() == 'EXCEL.EXE':
                    proc.kill()
                    sys.exit(0)

    def delete_all_files_in_directory(self, dir_path):
        for filename in os.listdir(dir_path):
            file_path = os.path.join(dir_path, filename)
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)

    def is_directory_empty(self, dir_path):
        return not bool(os.listdir(dir_path))

    def convert_xls_to_xlsx(self, folder_path, filename):
        hello = messagebox.askquestion('Найден файл .xls',
                                       'Найден файл .xls, напишите Да для попытки конвертации или напишите Нет для выхода',
                                       icon='warning')
        if hello == 'no':
            sys.exit(0)
        xls_file = pd.read_excel(os.path.join(folder_path, filename))
        xlsx_filename = filename.replace('.xls', '.xlsx')
        xls_file.to_excel(os.path.join(folder_path, xlsx_filename), index=False)
        os.remove(os.path.join(folder_path, filename))
        return xlsx_filename


app = App()
app.mainloop()
